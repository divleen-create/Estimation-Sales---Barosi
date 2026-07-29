"""Read-only parsing of the "Spend Split" workbook.

What the sheet holds: week-CUMULATIVE ad spend per platform, split by product
category (A2 Ghee / Cow Ghee / Others), with each category's share of that
platform's spend. Tabs are plain month names ("June", "July"). Layout per tab:

    row 1 |  1-7th July                       |  1-15th July      | 1-26th July
    row 2 |  Channels | A2 Ghee |  | Cow Ghee |  | Others |  | Total Spend | ...
    row 3+|  Blinkit  |  706390 | 90% | 45143 | 6% | 37441 | 4.75% | 788,974 | ...

So each "bucket" is a 7-column block: for every category a ₹ column and an
unlabelled % column, then Total Spend. A tab can carry any number of buckets and
the labels are read as written — nothing is assumed about the cadence.

Rules (from the brief):
  * show exactly what the sheet says; where a cell is missing or an error, N/A —
    never a guessed or filled-in number;
  * a bucket nobody has filled yet is marked pending (not zero);
  * months before config.SPEND_START_PERIOD are ignored (June);
  * this spend is NOT the same metric as the report's Ad Sales column (it is
    category-level ghee spend) — the two are never reconciled.

This module never writes to any source.
"""
from __future__ import annotations
import calendar
import datetime as _dt
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl

import config

SNAPSHOT = Path(__file__).resolve().parent / "snapshots" / "spend_snapshot.xlsx"

_MONTHS = {m.lower(): i for i, m in enumerate(calendar.month_name) if m}
_MONTHS.update({m.lower(): i for i, m in enumerate(calendar.month_abbr) if m})
_MONTHS["sept"] = 9
# "1-7th July", "1 - 15th", "1-21st Aug 26"
_BUCKET_RE = re.compile(r"1\s*[-–]\s*(\d{1,2})\s*(?:st|nd|rd|th)?", re.I)
_ERRORS = ("#DIV/0!", "#REF!", "#VALUE!", "#N/A", "#NAME?", "#NULL!", "#NUM!", "#ERROR!")


def _norm(v) -> str:
    return re.sub(r"\s+", " ", str(v)).strip().lower() if v is not None else ""


def _num(v) -> Optional[float]:
    """Cell -> number, or None when blank / an error / not numeric. Strings like
    '1,845,624' (text-entered) are accepted; error values are N/A, never 0."""
    if v is None or isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s or s.upper() in _ERRORS:
        return None
    s = s.replace(",", "").replace("₹", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _pct(v) -> Optional[float]:
    """Cell -> share as a fraction (0.90), or None for N/A. A percent-formatted
    cell reads back as 0.9; a text cell as '90%'; a plain 90 means 90%."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        if not s or s.upper() in _ERRORS:
            return None
        if s.endswith("%"):
            n = _num(s[:-1])
            return None if n is None else n / 100.0
    n = _num(v)
    if n is None:
        return None
    return n / 100.0 if n > 1.5 else n     # a share never exceeds 150%


# --- model ------------------------------------------------------------------
@dataclass
class SpendCell:
    spend: Optional[float] = None
    pct: Optional[float] = None


@dataclass
class SpendRow:
    channel: str
    cats: dict = field(default_factory=dict)      # category -> SpendCell
    total: Optional[float] = None                 # the sheet's own Total Spend
    @property
    def filled(self) -> bool:
        return any(c.spend is not None for c in self.cats.values()) or bool(self.total)


@dataclass
class SpendBucket:
    label: str                     # exactly as written, e.g. "1-26th July"
    end_day: Optional[int]         # 26
    rows: list = field(default_factory=list)
    @property
    def filled_rows(self) -> list:
        return [r for r in self.rows if r.filled]
    @property
    def pending(self) -> bool:
        return not self.filled_rows
    def cat_total(self, cat) -> Optional[float]:
        vals = [r.cats[cat].spend for r in self.filled_rows
                if cat in r.cats and r.cats[cat].spend is not None]
        return sum(vals) if vals else None
    def grand_total(self) -> Optional[float]:
        vals = [r.total for r in self.filled_rows if r.total is not None]
        return sum(vals) if vals else None


@dataclass
class SpendTable:
    month_label: str
    year: int
    month: int
    tab: str
    buckets: list = field(default_factory=list)
    channels: list = field(default_factory=list)
    issues: list = field(default_factory=list)   # sheet-quality notes for QC
    source: str = "Spend Split"
    @property
    def latest(self) -> Optional[SpendBucket]:
        """The most advanced bucket anyone has filled — what leadership reads first."""
        done = [b for b in self.buckets if not b.pending]
        return max(done, key=lambda b: (b.end_day or 0)) if done else None


# --- parsing ----------------------------------------------------------------
def _tab_period(name: str, default_year: int) -> Optional[tuple[int, int]]:
    """'July' -> (default_year, 7); 'July 26'/'Jul-2026' -> (2026, 7)."""
    n = _norm(name).replace("'", " ")
    m = re.match(r"^([a-z]+)\s*[-/]?\s*(\d{2,4})?$", n)
    if not m:
        return None
    mon = _MONTHS.get("sept" if m.group(1).startswith("sept") else m.group(1)[:3])
    if not mon:
        return None
    yr = m.group(2)
    if yr:
        yr = int(yr)
        yr += 2000 if yr < 100 else 0
    else:
        yr = default_year
    return (yr, mon)


def _header_rows(ws) -> tuple[int, int]:
    """(bucket-label row, column-header row). The header row is the one carrying
    'Channels'; bucket labels sit on the row above it (or on the same row)."""
    for r in range(1, min(8, ws.max_row) + 1):
        for c in range(1, min(4, ws.max_column) + 1):
            if "channel" in _norm(ws.cell(r, c).value):
                return (max(1, r - 1), r)
    return (1, 2)


def _buckets_of(ws, label_row: int, head_row: int) -> list[tuple[str, int, int]]:
    """[(label, first_col, last_col)] — one per cumulative block on the tab."""
    starts = []
    for c in range(1, ws.max_column + 1):
        for r in {label_row, head_row}:
            v = ws.cell(r, c).value
            if v is not None and _BUCKET_RE.search(str(v)):
                starts.append((c, re.sub(r"\s+", " ", str(v)).strip()))
                break
    out = []
    for i, (c, label) in enumerate(starts):
        end = (starts[i + 1][0] - 1) if i + 1 < len(starts) else ws.max_column
        out.append((label, c, end))
    return out


def _columns_of(ws, head_row: int, c0: int, c1: int) -> tuple[dict, Optional[int]]:
    """Within one block: {category: (spend_col, pct_col)} + the Total Spend col.
    The % column is the unlabelled one immediately right of each category."""
    cats, total_col = {}, None
    for c in range(c0, c1 + 1):
        h = _norm(ws.cell(head_row, c).value)
        if not h:
            continue
        if "total" in h:
            total_col = c
            continue
        for key, cat in config.SPEND_CATEGORY_KEYS.items():
            if key in h and cat not in cats:
                cats[cat] = (c, c + 1 if c + 1 <= c1 else None)
                break
    return cats, total_col


def parse_tab(path, tab: str, year: int, month: int, month_label: str) -> SpendTable:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[tab]
    label_row, head_row = _header_rows(ws)
    table = SpendTable(month_label=month_label, year=year, month=month, tab=tab)
    blocks = _buckets_of(ws, label_row, head_row)
    if not blocks:
        table.issues.append(f"{month_label}: no '1-Nth' cumulative block found on tab '{tab}'")
        return table

    # Channel rows: below the header, while column A holds a channel name. The tab
    # carries extra label rows under 'Channels' (a 'Metric' row whose cells read
    # 'Ad Spend%'), so those are skipped rather than treated as a channel.
    ch_rows = []
    for r in range(head_row + 1, ws.max_row + 1):
        name = ws.cell(r, 1).value
        n = _norm(name)
        if not n or n in ("metric", "metrics", "channels"):
            if ch_rows:
                break                       # blank line after the data = end
            continue                        # still in the header stack
        if n in ("total", "grand total", "sub total", "subtotal"):
            break
        ch_rows.append((r, re.sub(r"\s+", " ", str(name)).strip()))
    table.channels = [n for _, n in ch_rows]

    dim = calendar.monthrange(year, month)[1]
    for label, c0, c1 in blocks:
        cats, total_col = _columns_of(ws, head_row, c0, c1)
        m = _BUCKET_RE.search(label)
        end_day = int(m.group(1)) if m else None
        bucket = SpendBucket(label=label, end_day=end_day)
        missing = [c for c in config.SPEND_CATEGORIES if c not in cats]
        if missing:
            table.issues.append(f"{month_label} · {label}: no column found for "
                                f"{', '.join(missing)}")
        for r, name in ch_rows:
            row = SpendRow(channel=name)
            for cat in config.SPEND_CATEGORIES:
                sc, pc = cats.get(cat, (None, None))
                row.cats[cat] = SpendCell(
                    spend=_num(ws.cell(r, sc).value) if sc else None,
                    pct=_pct(ws.cell(r, pc).value) if pc else None)
            row.total = _num(ws.cell(r, total_col).value) if total_col else None
            if row.total == 0 and not any(c.spend for c in row.cats.values()):
                row.total = None            # an untouched block reads 0 -> pending
            bucket.rows.append(row)
        table.buckets.append(bucket)
        if end_day and end_day > dim:
            table.issues.append(f"{month_label} · {label}: ends on day {end_day} but "
                                f"{month_label} has {dim} days")

    _audit(table, dim)
    return table


def _audit(table: SpendTable, dim: int) -> None:
    """Sheet-quality notes — reported (QC advisories + the mail), never guessed at."""
    for b in table.buckets:
        if b.pending:
            table.issues.append(f"{table.month_label} · {b.label}: nothing filled in yet "
                                f"— shown as N/A")
            continue
        blank = [r.channel for r in b.rows if not r.filled]
        if blank:
            table.issues.append(f"{table.month_label} · {b.label}: no data for "
                                f"{', '.join(blank)} — shown as N/A")
        for r in b.filled_rows:
            parts = [c.spend for c in r.cats.values() if c.spend is not None]
            if r.total is not None and parts and abs(sum(parts) - r.total) > max(1.0, 0.01 * r.total):
                table.issues.append(
                    f"{table.month_label} · {b.label} · {r.channel}: categories add to "
                    f"{sum(parts):,.0f} but Total Spend says {r.total:,.0f}")
            shares = [c.pct for c in r.cats.values() if c.pct is not None]
            if len(shares) == len(config.SPEND_CATEGORIES) and abs(sum(shares) - 1) > 0.03:
                table.issues.append(
                    f"{table.month_label} · {b.label} · {r.channel}: shares add to "
                    f"{sum(shares)*100:.1f}%, not 100%")
            for cat, c in r.cats.items():
                if c.spend and not c.pct:
                    table.issues.append(
                        f"{table.month_label} · {b.label} · {r.channel}: {cat} has spend "
                        f"{c.spend:,.0f} but its share reads "
                        f"{'0%' if c.pct == 0 else 'blank'} — shown as written")
    # Cadence: from August the buckets should be 1-7 / 1-15 / 1-21 / 1-<month end>.
    if (table.year, table.month) > tuple(config.SPEND_START_PERIOD):
        want = [d if d is not None else dim for d in config.SPEND_BUCKET_DAYS]
        have = [b.end_day for b in table.buckets if b.end_day]
        odd = [d for d in have if d not in want]
        if odd:
            table.issues.append(f"{table.month_label}: bucket end-day(s) {odd} are outside the "
                                f"expected {want} cadence — shown as written")


def load_month(year: int, month: int, month_label: str) -> Optional[SpendTable]:
    """Spend Split for one month, or None when there is nothing to show:
    no snapshot, no matching tab, or a month before SPEND_START_PERIOD (June)."""
    if (year, month) < tuple(config.SPEND_START_PERIOD):
        return None
    if not SNAPSHOT.exists():
        return None
    try:
        wb = openpyxl.load_workbook(SNAPSHOT, data_only=True, read_only=True)
        names = list(wb.sheetnames)
        wb.close()
    except Exception:
        return None
    for name in names:
        if _tab_period(name, year) == (year, month):
            try:
                return parse_tab(SNAPSHOT, name, year, month, month_label)
            except Exception as e:  # noqa: BLE001 — never break the daily publish
                t = SpendTable(month_label=month_label, year=year, month=month, tab=name)
                t.issues.append(f"{month_label}: could not read tab '{name}' — {e}")
                return t
    return None


if __name__ == "__main__":
    import sys
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    t = load_month(2026, 7, "July 2026")
    if not t:
        raise SystemExit("no Spend Split data (snapshot missing or month before start)")
    print(f"{t.source} · {t.month_label} · tab '{t.tab}' · channels: {', '.join(t.channels)}")
    for b in t.buckets:
        print(f"\n== {b.label} (to day {b.end_day}) "
              f"{'PENDING' if b.pending else f'total {b.grand_total():,.0f}'}")
        for r in b.rows:
            bits = " | ".join(
                f"{c}: " + ("N/A" if r.cats[c].spend is None else f"{r.cats[c].spend:,.0f}")
                + ("" if r.cats[c].pct is None else f" ({r.cats[c].pct*100:.2f}%)")
                for c in config.SPEND_CATEGORIES)
            print(f"   {r.channel:18} {bits} | total "
                  f"{'N/A' if r.total is None else f'{r.total:,.0f}'}")
    print("\nissues:")
    for i in t.issues:
        print("  -", i)
