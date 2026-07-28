"""Read-only data access + parsing.

Reads the local .xlsx snapshots today. A Google Sheets reader (read-only
scope) can be dropped in behind `load_channels()` without changing anything
downstream. This module NEVER writes to any source.
"""
from __future__ import annotations
import calendar
import datetime as _dt
import os
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.utils import range_boundaries

import config

# Workbook load cache (keyed by path + mode + mtime) so repeated reads within a
# run don't re-parse the large workbooks. Read-only; never mutated.
_WB_CACHE: dict[tuple, tuple[float, object]] = {}


def _load_wb(path, data_only: bool = False, read_only: bool = False):
    key = (str(path), data_only, read_only)
    mt = os.path.getmtime(path)
    hit = _WB_CACHE.get(key)
    if hit and hit[0] == mt:
        return hit[1]
    wb = openpyxl.load_workbook(path, data_only=data_only, read_only=read_only)
    _WB_CACHE[key] = (mt, wb)
    return wb


@dataclass
class DailyRecord:
    date: _dt.date
    units: Optional[float] = None
    gmv: Optional[float] = None
    lm_gmv: Optional[float] = None
    ad_spend: Optional[float] = None


@dataclass
class Channel:
    name: str
    records: list[DailyRecord] = field(default_factory=list)


def _norm(s) -> str:
    """Lower-case, collapse whitespace, strip. Handles the stray leading
    spaces in headers like '                Shopify' and ' Minutes'."""
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s)).strip().lower()


def _group_map_from_row1(ws) -> dict[int, str]:
    """Map each column index -> group (channel) name, expanding merged cells
    in row 1 so every column under a merged header inherits its name."""
    col_group: dict[int, str] = {}
    # Merged ranges give the true span of each group header.
    for rng in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = range_boundaries(str(rng))
        if min_row == 1:
            name = ws.cell(1, min_col).value
            for c in range(min_col, max_col + 1):
                col_group[c] = name
    # Any non-merged, non-empty row-1 cells (single-column groups).
    for c in range(1, ws.max_column + 1):
        if c not in col_group:
            v = ws.cell(1, c).value
            if v not in (None, ""):
                col_group[c] = v
    # Forward-fill: a column with no row-1 header belongs to the channel to its
    # left. Fixes blocks whose merge is narrower than the block, e.g. Amazon's
    # header merges only G:I but its 'AD sales' column J is un-merged.
    last = None
    for c in range(1, ws.max_column + 1):
        val = col_group.get(c)
        if val not in (None, ""):
            last = val
        elif last is not None:
            col_group[c] = last
    return col_group


def _prev_month_names(ym) -> set:
    """Lower-case name + 3-letter abbrev of the month BEFORE ym, e.g.
    (2026, 3) -> {'february', 'feb'}. Used to spot a 'Feb GMV'-style LM header."""
    if not ym:
        return set()
    _, m = ym
    pm = 12 if m == 1 else m - 1
    full = calendar.month_name[pm].lower()
    return {full, full[:3]}


def _field_of(header: str, prev_names: set) -> Optional[str]:
    """Map a row-2 header to a canonical field, tolerant of month-labelled
    variants some tabs use (e.g. 'March GMV', 'Feb GMV', 'March Ad sales')
    instead of the standard 'Value' / 'LM GMV' / 'Ad sales'."""
    h = _norm(header)
    if not h:
        return None
    if "unit" in h:
        return "units"
    if "ad" in h:                                  # ad sales / ads sales / <month> ad sales
        return "ad_spend"
    if "gmv" in h or "value" in h:                  # 'Value', 'Total Value', 'March GMV'
        if "lm" in h or "last" in h or any(n in h for n in prev_names):
            return "lm_gmv"                         # 'LM GMV', 'Last Month GMV', prev-month GMV
        return "gmv"
    return config.FIELD_ALIASES.get(h)              # total sale/dolchi -> skipped by caller


def _column_map(ws, ym=None) -> dict[str, dict[str, int]]:
    """channel display name -> {field_name: column_index}, from rows 1-2.
    Skips Total/xx/excluded channels and keeps only units/gmv/lm_gmv/ad_spend.
    `ym` (year, month) lets month-labelled GMV headers be classified correctly."""
    col_group = _group_map_from_row1(ws)
    prev_names = _prev_month_names(ym)
    channels: dict[str, dict[str, int]] = {}
    for c in range(2, ws.max_column + 1):  # col 1 is DATE
        grp = _norm(col_group.get(c))
        if grp in config.CHANNEL_SKIP:
            continue
        display = config.CHANNEL_ALIASES.get(grp)
        if display is None or display in config.EXCLUDED_CHANNELS:
            continue
        fld = _field_of(ws.cell(2, c).value, prev_names)
        if fld in (None, "dolchi", "total_sale"):
            continue
        channels.setdefault(display, {})[fld] = c
    return channels


def _date_rows(ws) -> list[tuple[int, _dt.date]]:
    """Row index + date for each daily row (row 3 down while col A is a date)."""
    rows: list[tuple[int, _dt.date]] = []
    for r in range(3, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, _dt.datetime):
            rows.append((r, v.date()))
        elif isinstance(v, _dt.date):
            rows.append((r, v))
        elif rows:
            break  # first non-date after data begins => end of daily block
    return rows


def parse_tab(path, tab: str) -> list[Channel]:
    """Parse one workbook tab into a list of real Channels (skips Total/xx
    and excluded channels). Column mapping is derived from rows 1-2 so it
    survives layout differences between the two workbooks."""
    ws = _load_wb(path, data_only=True)[tab]
    channels = _column_map(ws, parse_period_from_name(tab))
    date_rows = _date_rows(ws)

    result: list[Channel] = []
    for display, fld_cols in channels.items():
        ch = Channel(name=display)
        for r, d in date_rows:
            rec = DailyRecord(date=d)
            for fld, c in fld_cols.items():
                val = ws.cell(r, c).value
                if isinstance(val, (int, float)):
                    setattr(rec, fld, float(val))
            ch.records.append(rec)
        result.append(ch)
    return result


@dataclass
class SheetChannelDiag:
    total_gmv: Optional[float] = None
    total_lm: Optional[float] = None
    total_units: Optional[float] = None
    total_ad: Optional[float] = None
    est_value: Optional[float] = None      # computed value in the estimate cell
    est_divisor: Optional[int] = None      # days the sheet divided by
    est_mult: Optional[int] = None         # days it projected to
    est_cell: Optional[str] = None
    ad_contri_excel: Optional[float] = None  # sheet's own Ad/Value ratio, if present


_EST_RE = re.compile(r"^=\s*[A-Za-z]+\d+\s*/\s*(\d+)\s*\*\s*(\d+)\s*$")
# Ad-contribution ratio the sheet sometimes stores in the ad column: '=M34/K34'.
_RATIO_RE = re.compile(r"^=[A-Za-z]+\d+/[A-Za-z]+\d+$")


def sheet_diagnostics(path, tab: str) -> dict[str, SheetChannelDiag]:
    """Read the sheet's OWN Total-row values and estimate-row formulas per
    channel — ground truth that QC compares against (independent of our math).
    The estimate divisor is parsed from the formula (e.g. '=H35/19*31' -> 19)
    so we can tell whether the sheet's day-count is stale."""
    wsv = _load_wb(path, data_only=True)[tab]
    wsf = _load_wb(path, data_only=False)[tab]
    ym = parse_period_from_name(tab)
    channels = _column_map(wsv, ym)
    rows = _date_rows(wsv)
    if not rows:
        return {}
    last = rows[-1][0]
    dim = calendar.monthrange(*ym)[1] if ym else 31  # days in month (estimate mult)

    # Totals row = the row just below the daily block with the most numeric
    # GMV cells across channels (Total/SUM row).
    best_row, best_hits = None, 0
    for r in range(last + 1, min(last + 8, wsv.max_row) + 1):
        hits = sum(1 for cols in channels.values()
                   if "gmv" in cols and isinstance(wsv.cell(r, cols["gmv"]).value, (int, float)))
        if hits > best_hits:
            best_hits, best_row = hits, r

    out: dict[str, SheetChannelDiag] = {}
    for name, cols in channels.items():
        d = SheetChannelDiag()
        if best_row and "gmv" in cols:
            d.total_gmv = wsv.cell(best_row, cols["gmv"]).value
            if "lm_gmv" in cols:
                d.total_lm = wsv.cell(best_row, cols["lm_gmv"]).value
            if "units" in cols:
                d.total_units = wsv.cell(best_row, cols["units"]).value
            if "ad_spend" in cols:
                d.total_ad = wsv.cell(best_row, cols["ad_spend"]).value
        # estimate: prefer the formula divisor (e.g. '=H35/19*31' -> 19); if the
        # snapshot has no formulas (e.g. a live Sheets-API pull), derive the
        # divisor from values instead: divisor = round(total_gmv * dim / est).
        if "gmv" in cols:
            gcol = cols["gmv"]
            for r in range(last + 1, min(last + 12, wsf.max_row) + 1):
                f = wsf.cell(r, gcol).value
                if isinstance(f, str):
                    m = _EST_RE.match(f.replace(" ", ""))
                    if m:
                        d.est_divisor, d.est_mult = int(m.group(1)), int(m.group(2))
                        d.est_value = wsv.cell(r, gcol).value
                        d.est_cell = wsv.cell(r, gcol).coordinate
                        break
            if d.est_divisor is None and d.total_gmv:
                for r in range(last + 1, min(last + 12, wsv.max_row) + 1):
                    v = wsv.cell(r, gcol).value
                    if isinstance(v, (int, float)) and v > d.total_gmv * 1.02:
                        d.est_value, d.est_mult = v, dim
                        d.est_divisor = round(d.total_gmv * dim / v)
                        d.est_cell = wsv.cell(r, gcol).coordinate
                        break
        # ad-contribution: the sheet's Ad/Value ratio, from formula ('=M34/K34')
        # or, in a values-only snapshot, a small ratio value in the ad column.
        if "ad_spend" in cols:
            acol = cols["ad_spend"]
            for r in range(last + 1, min(last + 6, wsf.max_row) + 1):
                f = wsf.cell(r, acol).value
                if isinstance(f, str) and _RATIO_RE.match(f.replace(" ", "")):
                    d.ad_contri_excel = wsv.cell(r, acol).value
                    break
            if d.ad_contri_excel is None:
                for r in range(last + 1, min(last + 6, wsv.max_row) + 1):
                    v = wsv.cell(r, acol).value
                    if isinstance(v, (int, float)) and 0 < v < 3:
                        d.ad_contri_excel = v
                        break
        out[name] = d
    return out


# --- Reporting-period resolution (auto-rolling month) -----------------------
_MONTHS = {m[:3].lower(): i for i, m in enumerate(calendar.month_name) if m}
_MONTHS["sept"] = 9  # sheets use both 'Sep' and 'Sept'


@dataclass
class Period:
    year: int
    month: int
    days_in_month: int
    label: str          # e.g. "July 2026"
    tab_qc: Optional[str]   # may be None for months only one sheet covers
    tab_mkt: Optional[str]
    pending_note: Optional[str] = None  # e.g. newer month exists but is empty
    conflicts: list = field(default_factory=list)  # cross-sheet GMV mismatches


def parse_period_from_name(name: str) -> Optional[tuple[int, int]]:
    """'July26'->(2026,7), 'FEB 26'->(2026,2), 'Sept25'->(2025,9),
    'July-2023'->(2023,7). Non-month tabs ('working','Trends') -> None."""
    n = re.sub(r"[\s\-']", "", str(name)).lower()
    m = re.match(r"^([a-z]+)(\d{2,4})$", n)
    if not m:
        return None
    mon_txt, yr_txt = m.group(1), m.group(2)
    mon = _MONTHS.get("sept" if mon_txt.startswith("sept") else mon_txt[:3])
    if not mon:
        return None
    yr = int(yr_txt)
    if yr < 100:
        yr += 2000
    return (yr, mon)


def _month_map(path) -> dict[tuple[int, int], str]:
    """Map (year, month) -> sheet/tab name for a workbook."""
    wb = _load_wb(path, data_only=True, read_only=True)
    out: dict[tuple[int, int], str] = {}
    for name in wb.sheetnames:
        ym = parse_period_from_name(name)
        if ym and ym not in out:
            out[ym] = name
    return out


def _tab_has_data(path, tab: str) -> bool:
    try:
        chans = parse_tab(path, tab)
    except Exception:
        return False
    return any(r.gmv is not None for ch in chans for r in ch.records)


def resolve_period(qc_path, mkt_path) -> Period:
    """Pick the latest month present AND populated in both workbooks.

    Rolls forward automatically (July -> August once August has data) and
    falls back to the last populated month if the newer tab is still empty.
    Set config.FORCE_PERIOD to pin a specific (year, month).
    """
    qc_map, mkt_map = _month_map(qc_path), _month_map(mkt_path)
    common = sorted(set(qc_map) & set(mkt_map), reverse=True)  # newest first
    if not common:
        raise RuntimeError("No month tab is present in both workbooks.")

    def make(ym: tuple[int, int], note: Optional[str] = None) -> Period:
        y, mth = ym
        return Period(y, mth, calendar.monthrange(y, mth)[1],
                      _dt.date(y, mth, 1).strftime("%B %Y"),
                      qc_map[ym], mkt_map[ym], note)

    if config.FORCE_PERIOD and tuple(config.FORCE_PERIOD) in set(common):
        return make(tuple(config.FORCE_PERIOD))

    populated = [ym for ym in common
                 if _tab_has_data(qc_path, qc_map[ym]) and _tab_has_data(mkt_path, mkt_map[ym])]
    if not populated:
        # Nothing populated yet -> show the newest common month as-is (fallback).
        return make(common[0], "No data populated yet — showing the latest tab as-is.")

    chosen = populated[0]
    note = None
    if common[0] != chosen:  # a newer month tab exists but isn't populated yet
        ny, nm = common[0]
        note = (f"{_dt.date(ny, nm, 1).strftime('%B %Y')} is not populated yet — "
                f"showing {make(chosen).label} (last data available).")
    return make(chosen, note)


# --- Loading ---------------------------------------------------------------
def _sources() -> tuple[Path, Path]:
    """Prefer read-only Google Sheets snapshots when present, else local .xlsx."""
    snap = Path(__file__).resolve().parent / "snapshots"
    qc_src, mkt_src = snap / "qc_snapshot.xlsx", snap / "mkt_snapshot.xlsx"
    if qc_src.exists() and mkt_src.exists():
        return qc_src, mkt_src
    return config.WORKBOOK_QC, config.WORKBOOK_MKT


def list_periods() -> list[Period]:
    """Every month present in EITHER workbook (union), newest first, that has
    data in whichever sheet(s) contain it.

    A month may live in only one sheet — e.g. the Quick Commerce sheet has
    history back to 2023 while the Marketplace & D2C sheet starts later. Such a
    month is included and renders only the section(s) that have data. Ignores
    config.FORCE_PERIOD (that only pins the single 'current' month).
    """
    qc_path, mkt_path = _sources()
    qc_map, mkt_map = _month_map(qc_path), _month_map(mkt_path)
    out: list[Period] = []
    for ym in sorted(set(qc_map) | set(mkt_map), reverse=True):  # newest first
        qtab, mtab = qc_map.get(ym), mkt_map.get(ym)
        q_ok = bool(qtab) and _tab_has_data(qc_path, qtab)
        m_ok = bool(mtab) and _tab_has_data(mkt_path, mtab)
        if not (q_ok or m_ok):
            continue
        y, mth = ym
        out.append(Period(y, mth, calendar.monthrange(y, mth)[1],
                          _dt.date(y, mth, 1).strftime("%B %Y"),
                          qtab if q_ok else None, mtab if m_ok else None))
    return out


def _channel_gmv(ch: "Channel") -> float:
    return sum(r.gmv for r in ch.records if r.gmv is not None)


def _has_gmv(ch: "Channel") -> bool:
    return any(r.gmv is not None for r in ch.records)


def _merge_channels(qc_list, mkt_list, label):
    """Merge the two sheets for ONE month by OVERWRITE (never sum): one series
    per channel, preferring the sheet that actually has GMV data. When BOTH
    sheets have data for a channel but the totals disagree, record a conflict
    (for an email alert) and keep the richer series."""
    qc = {c.name: c for c in qc_list}
    mkt = {c.name: c for c in mkt_list}
    order = list(qc.keys()) + [n for n in mkt if n not in qc]
    merged, conflicts = [], []
    for name in order:
        a, b = qc.get(name), mkt.get(name)
        if a and b:
            ad, bd = _has_gmv(a), _has_gmv(b)
            if ad and bd:
                ga, gb = _channel_gmv(a), _channel_gmv(b)
                if abs(ga - gb) > max(1.0, 0.001 * max(abs(ga), abs(gb))):
                    # Sheets disagree — do NOT guess. Withhold this channel for the
                    # month (left blank in the output) and alert; user reconciles
                    # the sheets and re-runs. (Dropping keeps QC clean.)
                    conflicts.append(f"{label} · {name}: QC ₹{ga:,.0f} vs MKT ₹{gb:,.0f}")
                else:
                    merged.append(b)          # values agree — keep either
            elif ad:
                merged.append(a)              # only QC has data
            elif bd:
                merged.append(b)              # only MKT has data
            # neither sheet has GMV for this channel this month -> skip
        elif a or b:
            merged.append(a or b)             # present in just one sheet
    return merged, conflicts


def load_channels(period: "Period | None" = None) -> tuple[list[Channel], Period]:
    """Return (channels, reporting Period). Both sheets are read for the month
    and merged by OVERWRITE (prefer the sheet with data; never sum). If `period`
    is given, load exactly its tabs; else auto-resolve to the latest month."""
    qc_src, mkt_src = _sources()
    period = period or resolve_period(qc_src, mkt_src)
    qc_list = parse_tab(qc_src, period.tab_qc) if period.tab_qc else []
    mkt_list = parse_tab(mkt_src, period.tab_mkt) if period.tab_mkt else []
    chans, conflicts = _merge_channels(qc_list, mkt_list, period.label)
    period.conflicts = conflicts
    return chans, period


def load_sheet_diagnostics(period: "Period | None" = None) -> dict[str, SheetChannelDiag]:
    """Sheet Total-row values + estimate formulas for every channel, for the
    given period (or the resolved latest). Used by QC as ground truth."""
    qc_src, mkt_src = _sources()
    period = period or resolve_period(qc_src, mkt_src)
    out: dict[str, SheetChannelDiag] = {}
    if period.tab_qc:
        out.update(sheet_diagnostics(qc_src, period.tab_qc))
    if period.tab_mkt:
        out.update(sheet_diagnostics(mkt_src, period.tab_mkt))
    return out


if __name__ == "__main__":
    chans, period = load_channels()
    print(f"Period: {period.label} | tabs: {period.tab_qc} / {period.tab_mkt}"
          f" | {period.days_in_month} days")
    if period.pending_note:
        print("Note:", period.pending_note)
    for ch in chans:
        days = [r for r in ch.records if r.gmv is not None]
        last = max((r.date for r in days), default=None)
        total = sum(r.gmv or 0 for r in ch.records)
        print(f"  {ch.name:18} days_with_gmv={len(days):2}  last={last}  MTD_gmv={total:,.0f}")
