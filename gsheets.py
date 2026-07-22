"""READ-ONLY live pull from Google Sheets via the Sheets API.

Why the Sheets API (not Drive export): these sheets have "downloading/exporting
disabled for viewers", which blocks the Drive xlsx/CSV export even for an
authenticated viewer. The Sheets API `values.get` is a data read that a Viewer
service account CAN use, so it is the reliable read-only path. It returns
computed values (SUM totals, estimates, growth included), which is all the
pipeline needs — the estimate day-count is derived from those values in
data_source.sheet_diagnostics.

It writes local .xlsx snapshots (report_tool/snapshots/) that the rest of the
pipeline reads unchanged. It NEVER writes to the Google Sheets.

SETUP (one time)
----------------
1. pip install google-api-python-client google-auth
2. Google Cloud console: create a project, enable the **Google Sheets API**.
3. Create a **service account**, add a **JSON key**, download it.
4. **Share BOTH sheets** with the service account's email (…@…iam.gserviceaccount.com)
   as **Viewer** (Share dialog, paste the email, Viewer, Send).
5. Point the tool at the key:  set env var GOOGLE_APPLICATION_CREDENTIALS to its path
   (PowerShell:  $env:GOOGLE_APPLICATION_CREDENTIALS="C:\\path\\to\\key.json")
6. Pull + build:  python -c "import gsheets; gsheets.fetch_snapshots()"  then  python main.py

The scope is fixed to spreadsheets.readonly — there is no code path that writes.
"""
from __future__ import annotations
import datetime as _dt
import os
from pathlib import Path

import openpyxl

import config
from data_source import parse_period_from_name

SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
SNAPSHOT_DIR = Path(__file__).resolve().parent / "snapshots"
_SHEETS = {
    config.GSHEET_QC_ID: SNAPSHOT_DIR / "qc_snapshot.xlsx",
    config.GSHEET_MKT_ID: SNAPSHOT_DIR / "mkt_snapshot.xlsx",
}
_RECENT_TABS = 4          # how many recent month tabs to snapshot per workbook
_EXCEL_EPOCH = _dt.datetime(1899, 12, 30)


def _credentials():
    from google.oauth2 import service_account
    key = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
    if not (key and Path(key).exists()):
        raise RuntimeError(
            "Set GOOGLE_APPLICATION_CREDENTIALS to your service-account key JSON "
            "(see gsheets.py header for setup).")
    return service_account.Credentials.from_service_account_file(key, scopes=SCOPES)


def _service():
    from googleapiclient.discovery import build
    return build("sheets", "v4", credentials=_credentials(), cache_discovery=False)


def _serial_to_dt(v):
    """Sheets API SERIAL_NUMBER date -> datetime (col A only)."""
    return _EXCEL_EPOCH + _dt.timedelta(days=float(v))


def _write_snapshot(dest: Path, tabs: list[tuple[str, list[list]]]) -> None:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for title, values in tabs:
        ws = wb.create_sheet(title=title[:31])  # Excel tab-name limit
        for r, row in enumerate(values, start=1):
            for c, val in enumerate(row, start=1):
                if c == 1 and isinstance(val, (int, float)) and val > 40000:
                    val = _serial_to_dt(val)  # date column
                ws.cell(r, c).value = val
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def fetch_snapshots() -> dict[str, Path]:
    """Pull the recent month tabs of both sheets (read-only) into local .xlsx
    snapshots. Returns {sheet_id: snapshot_path}."""
    svc = _service()
    for sheet_id, dest in _SHEETS.items():
        meta = svc.spreadsheets().get(spreadsheetId=sheet_id).execute()
        titles = [s["properties"]["title"] for s in meta["sheets"]]
        months = sorted(((parse_period_from_name(t), t) for t in titles),
                        key=lambda x: (x[0] or (0, 0)), reverse=True)
        month_titles = [t for ym, t in months if ym][:_RECENT_TABS]
        tabs = []
        for t in month_titles:
            resp = svc.spreadsheets().values().get(
                spreadsheetId=sheet_id, range=f"'{t}'",
                valueRenderOption="UNFORMATTED_VALUE",
                dateTimeRenderOption="SERIAL_NUMBER").execute()
            tabs.append((t, resp.get("values", [])))
        _write_snapshot(dest, tabs)
        print(f"snapshot -> {dest}  ({', '.join(month_titles)})")
    return dict(_SHEETS)


if __name__ == "__main__":
    fetch_snapshots()
